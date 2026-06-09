# converts the assets.json file to an Excel spreadsheet with formatting, comments, and hyperlinks into human readable format. 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import json
from pathlib import Path
from datetime import datetime

assets_json = Path(__file__).resolve().parents[1] / 'assets' / 'data' / 'assets.json'

with open(assets_json, encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = 'Asset Table'

COLUMN_META = {
    "Asset ID": {"comment": "Unique identifier for the asset, e.g., 'A[0-9]{3}'"},
    "Asset Name": {"comment": "Name of the asset"},
    "Asset Type": {"comment": "Type of the asset e.g., photo | illustration"},
    "Platform": {"comment": "Platform where the asset is hosted e.g. Pixabay | Unsplash | Pexels"},
    "Source URL": {"comment": "URL of the preview page for the asset"},
    "Licence": {"comment": "Licence information for the asset (hyperlinked)"},
    "Author": {"comment": "Author of the asset (hyperlinked)"},
    "Original Format": {"comment": "Original format of the asset e.g., JPEG, PNG, JPG"},
    "Dimensions": {"comment": "Dimensions of the asset (e.g., 1920 x 1080)"},
    "Aspect Ratio": {"comment": "Aspect ratio of the asset (e.g., 16:9)"},
    "File Name": {"comment": "Name of the original file"},
    "File Format": {"comment": "Format of the file (e.g., JPEG, PNG) - this will be editied to match the modified file format if the asset is optimised"},
    "File Size": {"comment": "Size of the file in bytes - this will be edited to match the modified file size if the asset is optimised"},
    "Alt Text": {"comment": "Alternative text for the asset (e.g., 'A beautiful landscape')"},
    "Date Published": {"comment": "Date when the asset was published (e.g., 2021-02-16T13:01:07Z)"},
    "Date Accessed": {"comment": "Date when the asset was accessed (e.g., 2021-02-16T13:01:07Z)"},
    "Notes / Optimisation": {"comment": "Notes or optimisation information for the asset"},
    "Direct Image URL": {"comment": "Direct URL to the image file (not the preview page)"},
    "Platform ID": {"comment": "Identifier for the platform the asset is hosted on"},
    "Local Image Path": {"comment": "Local path to the image file"},
    "Modifications": {"comment": "Details of any modifications made to the original asset (e.g., 'Cropped to 16:9 aspect ratio and optimized for web')"}
}

# Create header row with the keys from COLUMN_META 
headers = list(COLUMN_META.keys())

ws.append(headers)

for i in range(1, len(headers) + 1):
    # Make header bold and add comment if available
    ws.cell(row=1, column=i).font = Font(bold=True)
    if headers[i - 1] in COLUMN_META:
        comment = COLUMN_META[headers[i - 1]].get("comment")
        if comment: ws.cell(row=1, column=i).comment = Comment(comment, "System")


def link_value(value, default_name='N/A', default_url=''):
    # If the value is a dict with 'name' and 'url', create a hyperlink formula
    if isinstance(value, dict):
        name = value.get('name') or default_name
        url = value.get('url') or default_url
        if url: return f'=HYPERLINK("{url}","{name}")'
        return name
    if value: return value
    return default_name


def get_field(obj, *keys):
    # Try each key in order, checking for both original and lowercase versions
    if not isinstance(obj, dict): return None
    for k in keys:
        if k in obj and obj[k] is not None: return obj[k]
    for k in keys:
        alt = k.lower()
        if alt in obj and obj[alt] is not None: return obj[alt]
    return None


def get_dimensions(asset):
    # Try 'originalResolution' first, then 'dimensions'
    return get_field(asset, 'originalResolution', 'Original Resolution', 'Resolution', 'original_resolution') or 'N/A'

def iso8601_to_date(timestamp):
    # e.g. 2021-02-16T13:01:07Z to 16 February 2021 at 13:01:07 UTC
    if not timestamp: return 'N/A'
    try:
        dt = datetime.strptime(timestamp, '%Y-%m-%dT%H:%M:%SZ')
        return dt.strftime('%d %B %Y at %H:%M:%S UTC')
    except ValueError:
        return timestamp

for a in data.get('assets', []):
    ws.append([
        get_field(a, 'assetId') or 'N/A',
        get_field(a, 'assetName') or 'N/A',
        get_field(a, 'assetType') or 'N/A',
        get_field(a, 'platform') or 'N/A',
        get_field(a, 'sourceUrl') or 'N/A',
        link_value(get_field(a, 'license')),
        link_value(get_field(a, 'author')),
        get_field(a, 'originalFormat') or 'N/A',
        get_dimensions(a),
        get_field(a, 'aspectRatio') or 'N/A',
        get_field(a, 'fileName') or 'N/A',
        get_field(a, 'fileFormat') or 'N/A',
        get_field(a, 'fileSize') or 'N/A',
        get_field(a, 'altText') or 'N/A',
        iso8601_to_date(get_field(a, 'datePublished')),
        get_field(a, 'dateAccessed') or 'N/A', 
        get_field(a, 'notes') or 'N/A',
        get_field(a, 'directImageUrl') or 'N/A',
        get_field(a, 'platformId') or 'N/A',
        get_field(a, 'localImagePath') or 'N/A',
        get_field(a, 'modifications') or 'Optimised via https://shortpixel.com/online-image-compression batch (lossy compression), converted all PNG and JPEG into JPG format, Made sure most JS Files are under 1MB and all images are under 500KB for optimal web performance'
    ])


ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
t = Table(displayName='AssetTable', ref=ref)
t.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
ws.add_table(t)

widths = [12, 25, 15, 15, 45, 25, 30, 15, 20, 15, 25, 15, 12, 45, 18, 15, 45, 45, 18, 30, 45]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

nowrap = [1, 13]
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        if cell.column not in nowrap:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

ws.freeze_panes = 'B2' # makes header and first column static for easier navigation when scrolling through large datasets (Asset ID + header row will be frozen)

file = str(Path(__file__).resolve().parents[1] / 'assets' / 'data' / 'assets.xlsx')
wb.save(file)
print(f"Excel file saved: {file}")